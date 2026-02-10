import os
import msvcrt
from openpyxl import load_workbook
from openpyxl.styles import Font

def print_progress(current, total, bar_length=30):
    percent = current / total if total else 0
    filled = int(bar_length * percent)
    bar = "█" * filled + "-" * (bar_length - filled)
    print(f"\rProgress: |{bar}| {percent*100:5.1f}% ({current}/{total})", end="")

def write_excel_formulas(folder_path, l_value):
    folder_path = folder_path.strip().strip('"')
    xlsx_files = [os.path.join(root, file) 
                  for root, _, files in os.walk(folder_path) 
                  for file in files 
                  if file.endswith(".xlsx") and not file.startswith("~$")]

    total = len(xlsx_files)
    if total == 0:
        print("⚠️ No Excel files found.")
        return

    processed = 0
    failed = 0

    for i, path in enumerate(xlsx_files, start=1):
        try:
            wb = load_workbook(path)
            sheet_names = wb.sheetnames

            # 1. 첫 번째 시트의 B5 셀 볼드 처리
            if sheet_names:
                ws1 = wb[sheet_names[0]]
                ws1["B5"].font = Font(bold=True)

            # 2. 세 번째 시트 작업 (기존 로직)
            if len(sheet_names) < 3:
                raise Exception("시트가 3개 미만입니다.")
            
            ws3 = wb[sheet_names[2]]
            last_row = ws3.max_row

            ws3.cell(row=1, column=5).value = "resistivity (ohm/cm)"
            ws3.cell(row=1, column=6).value = "conductivity (S/cm)"

            for r in range(2, last_row + 1):
                ws3.cell(row=r, column=5).value = f"=(D{r}/C{r})*(2*PI()*{l_value})"
                ws3.cell(row=r, column=6).value = f"=1/E{r}"

            # H1, H2 평균값 및 볼드 설정
            ws3["H1"].value = "Average Conductivity"
            ws3["H1"].font = Font(bold=True)
            ws3["H2"].value = f"=AVERAGE(F3:F{last_row})"
            ws3["H2"].font = Font(bold=True)

            wb.save(path)
            processed += 1

        except Exception as e:
            failed += 1
            print(f"\n❌ Failed: {path}\n   Reason: {e}")

        print_progress(i, total)

    print("\n\n✅ Processing completed!")
    print(f"✅ Processed: {processed} | ❌ Failed: {failed}\n")

if __name__ == "__main__":
    while True:
        folder = input("📂 Enter folder path: ").strip()
        try:
            l_input = float(input("✍️ Enter l (cm): "))
        except ValueError:
            print("❌ 숫자를 입력해주세요.")
            continue

        write_excel_formulas(folder, l_input)

        print("Press Ctrl+F to continue, any other key to exit")
        if msvcrt.getch() != b'\x06':
            break