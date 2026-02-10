import os
import msvcrt
from openpyxl import load_workbook
from openpyxl.styles import Font

def bold_b5_in_xlsx(folder_path):
    folder_path = folder_path.strip()
    if not os.path.isdir(folder_path):
        print("🚫 Invalid folder path.")
        return

    print("🔄 Processing Excel files (Setting B5 to Bold)...\n")

    # .xlsx 파일 수집 (하위 폴더 포함)
    xlsx_files = []
    for root_dir, _, files in os.walk(folder_path):
        for filename in files:
            # 임시 파일(~$...)은 제외하고 .xlsx만 수집
            if filename.endswith(".xlsx") and not filename.startswith("~$"):
                xlsx_files.append(os.path.join(root_dir, filename))

    total = len(xlsx_files)
    if total == 0:
        print("⚠️ No Excel files found.")
        return

    success = 0
    failed = 0

    # 파일 처리
    for i, xlsx_path in enumerate(xlsx_files, start=1):
        try:
            # 엑셀 파일 로드
            wb = load_workbook(xlsx_path)
            
            # 모든 시트의 B5 셀을 굵게 만들고 싶다면 아래 주석을 해제하세요.
            # 현재는 '첫 번째 시트'만 처리하도록 설정했습니다.
            ws = wb.active 
            
            # B5 셀 선택 및 폰트 설정 (Bold=True)
            target_cell = ws['B5']
            target_cell.font = Font(bold=True)

            # 저장
            wb.save(xlsx_path)
            success += 1
        except Exception as e:
            # print(f"\nError in {xlsx_path}: {e}") # 에러 원인 확인 필요 시 주석 해제
            failed += 1

        # 진행률 표시
        progress = (i / total) * 100
        print(f"Progress: {progress:.1f}% ({i}/{total})", end="\r")

    # 최종 결과 출력
    print("\n\n✅ Task completed!")
    print(f"Results:\n✨ Bold applied: {success}\n❌ Failed: {failed}\n")

if __name__ == "__main__":
    while True:
        folder_path = input("Enter the folder path containing Excel files: ")
        bold_b5_in_xlsx(folder_path)

        print("Press Ctrl+F to process more files, or any other key to exit...")
        key = msvcrt.getch()
        # Ctrl+F ASCII 코드: 6 (사용자 코드 주석에는 7이라 되어있으나 F는 6입니다)
        if key != b'\x06':
            break
        print("\n🔄 Restarting...\n")