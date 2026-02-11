import os
import re
import time
import msvcrt
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.styles import Font

# --- [1. 개별 파일 수식 및 서식 설정] ---
def print_progress(current, total, bar_length=30):
    """콘솔에 작업 진행률 표시바를 출력하는 함수"""
    percent = current / total if total else 0
    filled = int(bar_length * percent)
    bar = "█" * filled + "-" * (bar_length - filled)
    print(f"\r    Progress: |{bar}| {percent*100:5.1f}% ({current}/{total})", end="")

def apply_formulas_to_all_files(root_path, l_value):
    """지정된 경로 내의 모든 개별 데이터 파일에 전도도 계산 수식을 적용하는 함수"""
    xlsx_files = []
    # 폴더 내 모든 .xlsx 파일을 탐색 (임시 파일 및 Summary 파일 제외)
    for root, _, files in os.walk(root_path):
        for file in files:
            if file.endswith(".xlsx") and not file.startswith("~$") and "Summary" not in file:
                xlsx_files.append(os.path.join(root, file))
    
    total = len(xlsx_files)
    if total == 0: return
    
    print(f"📝 Step 1: Applying formulas to {total} files...")
    for i, path in enumerate(xlsx_files, start=1):
        try:
            wb = load_workbook(path)
            # 첫 번째 시트 서식 설정
            if wb.sheetnames:
                wb[wb.sheetnames[0]]["B5"].font = Font(bold=True)
            # 세 번째 시트에 resistivity 및 conductivity, average conductivity 수식 추가
            if len(wb.sheetnames) >= 3:
                ws3 = wb[wb.sheetnames[2]]
                last_row = ws3.max_row
                ws3.cell(row=1, column=5).value = "resistivity (ohm/cm)"
                ws3.cell(row=1, column=6).value = "conductivity (S/cm)"
                for r in range(2, last_row + 1):
                    ws3.cell(row=r, column=5).value = f"=(D{r}/C{r})*(2*PI()*{l_value})"
                    ws3.cell(row=r, column=6).value = f"=1/E{r}"
                ws3["H1"].value, ws3["H1"].font = "Average Conductivity", Font(bold=True)
                ws3["H2"].value, ws3["H2"].font = f"=AVERAGE(F3:F{last_row})", Font(bold=True)
            wb.save(path)
        except Exception as e:
            print(f"\n   ❌ Error in {os.path.basename(path)}: {e}")
        print_progress(i, total)
    print("\n   ✅ Step 1 Complete.")

# --- [2. 요약 파일 및 차트(Slope 추출) 생성] ---
def process_summary_folder(excel, folder_path):
    """폴더별 데이터를 취합하여 요약 엑셀 및 아레니우스/습도 차트를 생성하는 함수"""
    path_lower = folder_path.lower()
    # 폴더명에 따른 작업 유형 판별
    is_arrhenius = "arrhenius plot" in path_lower
    is_humidity = "humidity" in path_lower
    if not (is_arrhenius or is_humidity): return

    data_list = []
    # 폴더 내 파일들로부터 온도(oC) 또는 습도(%) 값을 추출하여 리스트화
    for root, _, files in os.walk(folder_path):
        for file in files:
            if file.endswith(".xlsx") and not file.startswith("~$") and "Summary" not in file:
                f_name = os.path.basename(root)
                num = 0
                if is_arrhenius:
                    m = re.search(r'(\d+)\s*oc', f_name, re.IGNORECASE)
                    num = int(m.group(1)) if m else 0
                else:
                    m = re.search(r'(\d+)\s*%', f_name)
                    num = int(m.group(1)) if m else 0
                data_list.append((num, os.path.join(root, file)))

    if not data_list: return
    # 온도/습도 순으로 정렬
    data_list.sort(key=lambda x: x[0])
    
    print(f"📊 Step 2: Creating summary for {os.path.basename(folder_path)}...")
    extracted = []
    # 각 파일에서 계산된 평균 전도도(H2 셀) 값을 추출
    for num, path in data_list:
        try:
            wb = excel.Workbooks.Open(os.path.abspath(path))
            val = wb.Sheets(3).Range("H2").Value
            extracted.append((num, val))
            wb.Close(False)
        except: pass
        
    if extracted:
        # 새로운 요약 엑셀 파일 생성
        wb_out = excel.Workbooks.Add()
        ws_out = wb_out.ActiveSheet
        last_row_idx = len(extracted) + 1
        
        # 아레니우스 플롯용 데이터 변환 (온도 -> 절대온도 -> 1000/T -> ln전도도)
        if is_arrhenius:
            for c, t in enumerate(["temp (℃)", "|temp| (K)", "1000/T", "σ (S/cm)", "ln(σ) (S/cm)"], 1): ws_out.Cells(1, c).Value = t
            for idx, (temp, sigma) in enumerate(extracted, 2):
                ws_out.Cells(idx, 1).Value = temp
                ws_out.Cells(idx, 2).Formula = f"=A{idx}+273.15"
                ws_out.Cells(idx, 3).Formula = f"=1000/B{idx}"
                ws_out.Cells(idx, 4).Value = sigma
                if sigma and sigma > 0: ws_out.Cells(idx, 5).Formula = f"=LN(D{idx})"
            last_c, x_t, y_t = "E", "1000/T (1000/K)", "ln (σ)"
            chart_range = ws_out.Range(f"C1:C{last_row_idx},E1:E{last_row_idx}")
        # 습도 요약 데이터 작성
        else:
            ws_out.Cells(1,1).Value, ws_out.Cells(1,2).Value = "RH (%)", "σ (S/cm)"
            for idx, (rh, sigma) in enumerate(extracted, 2):
                ws_out.Cells(idx,1).Value, ws_out.Cells(idx,2).Value = rh, sigma
            last_c, x_t, y_t = "B", "RH (%)", "σ (S/cm)"
            chart_range = ws_out.Range(ws_out.Cells(1,1), ws_out.Cells(last_row_idx, 2))

        # 표 헤더 및 정렬 설정
        ws_out.Rows(1).Font.Bold = True
        ws_out.Columns(f"A:{last_c}").HorizontalAlignment = -4108
        ws_out.Columns(f"A:{last_c}").AutoFit()

        # 분산형 차트 생성
        chart_obj = ws_out.ChartObjects().Add(60, 180, 450, 300)
        chart = chart_obj.Chart
        chart.ChartType = 74 # xlXYScatterLines (직선 및 표식이 있는 분산형)
        chart.SetSourceData(chart_range)
        chart.HasLegend = False
        
        # --- [차트 서식 설정: 진한 청록 강조 1 적용] ---
        if chart.SeriesCollection().Count > 0:
            series = chart.SeriesCollection(1)
            accent_color = 8544277 # RGB(21, 96, 130) - 색상 코드 공식 (B*65536) + (G*256) + R
            
            # 표식 설정 (원형, 크기 5)
            series.MarkerStyle = 8 
            series.MarkerSize = 5
            series.MarkerBackgroundColor = accent_color
            series.MarkerForegroundColor = accent_color
            
            # 선 설정 (너비 1.5pt)
            series.Format.Line.Weight = 1.5
            series.Format.Line.ForeColor.RGB = accent_color
        
        slope = 0
        # 아레니우스 플롯일 경우 기울기(Slope) 및 활성화 에너지 계산
        if is_arrhenius:
            y_range = f"E2:E{last_row_idx}"
            x_range = f"C2:C{last_row_idx}"

            ws_out.Range("H1").Value, ws_out.Range("H1").Font.Bold = "kB", True
            ws_out.Range("H2").Value = -0.086173
            ws_out.Range("H4").Value, ws_out.Range("H4").Font.Bold = "Slope", True
            # LINEST 함수를 이용해 선형 회귀 기울기 추출
            ws_out.Range("H5").Formula = f"=LINEST({y_range}, {x_range})"
            time.sleep(1)  # time sleep to ensure formula calculation
            slope = ws_out.Range("H5").Value or 0
            # 활성화 에너지(eV) 계산
            ws_out.Range("H7").Value, ws_out.Range("H7").Font.Bold = "Activation energy (eV)", True
            ws_out.Range("H8").Formula = "=H5*H2"
            
            # 차트에 선형 추세선 추가 및 수식 표시
            tl = series.Trendlines().Add(Type=-4132) # xlLinear
            tl.DisplayEquation = True
            ws_out.Columns("H").HorizontalAlignment = -4108
            ws_out.Columns("H").AutoFit()

        # 축 제목 및 서식 설정
        for ax_t, t_txt in zip([1, 2], [x_t, y_t]):
            ax = chart.Axes(ax_t)
            ax.HasTitle, ax.AxisTitle.Text = True, t_txt
            ax.AxisTitle.Font.Size, ax.AxisTitle.Font.Bold = 14, True
            ax.HasMajorGridlines = False

        # 요약 파일 저장 및 닫기
        save_path = os.path.join(folder_path, f"Summary_{os.path.basename(folder_path)}.xlsx")
        wb_out.SaveAs(os.path.abspath(save_path))
        wb_out.Close()
        print(f"      ✅ Summary created. (Slope: {slope})")

# --- [3. 메인 실행 루프] ---
def main():
    """사용자로부터 경로와 변수를 입력받아 전체 프로세스를 실행하는 메인 함수"""
    while True:
        root_input = input("\n📂 Enter root folder: ").strip().strip('"')
        if not os.path.isdir(root_input):
            print("❌ Invalid path."); continue
        # 시편 두께(l) 입력
        try:
            l_val = float(input("✍️ Enter l (cm): "))
        except:
            print("❌ Enter a number."); continue
        
        start = time.time()
        # 1단계: 개별 파일 수식 입력 실행
        apply_formulas_to_all_files(root_input, l_val)
        
        # 엑셀 어플리케이션 백그라운드 실행
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        excel.Visible, excel.DisplayAlerts = False, False
        try:
            # 2단계: 하위 폴더별 요약 작업 실행
            subs = [os.path.join(root_input, f) for f in os.listdir(root_input) if os.path.isdir(os.path.join(root_input, f))]
            for s in subs: process_summary_folder(excel, s)
        finally:
            # 작업 완료 후 엑셀 종료
            excel.Quit()
        
        print(f"\n✨ Done! ({time.time()-start:.1f}s)")
        print("Press [Ctrl+F] to continue, any other key to exit.")
        # Ctrl+F 입력 시 반복 실행, 그 외 종료
        if msvcrt.getch() != b'\x06': break

if __name__ == "__main__":
    main()