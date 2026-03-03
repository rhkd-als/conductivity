import os
import re
import time
import msvcrt
import win32com.client as win32
from openpyxl import load_workbook
from openpyxl.styles import Font

# --- [1. 개별 파일 수식 설정] ---
def print_progress(current, total, bar_length=30):
    percent = current / total if total else 0
    filled = int(bar_length * percent)
    bar = "█" * filled + "-" * (bar_length - filled)
    print(f"\r    Progress: |{bar}| {percent*100:5.1f}% ({current}/{total})", end="")

def apply_formulas_to_all_files(root_path, l_value):
    xlsx_files = []
    for root, _, files in os.walk(root_path):
        for file in files:
            if file.endswith(".xlsx") and not file.startswith("~$") and "Summary" not in file:
                xlsx_files.append(os.path.join(root, file))
    
    total = len(xlsx_files)
    if total == 0: return
    
    print(f"📝 Step 1: Applying formulas to {total} files...")
    for i, path in enumerate(xlsx_files, start=1):
        try:
            wb = load_workbook(path)
            if len(wb.sheetnames) >= 3:
                ws3 = wb[wb.sheetnames[2]]
                last_row = ws3.max_row
                ws3.cell(row=1, column=5).value = "resistivity (ohm/cm)"
                ws3.cell(row=1, column=6).value = "conductivity (S/cm)"
                for r in range(2, last_row + 1):
                    ws3.cell(row=r, column=5).value = f"=(D{r}/C{r})*(2*PI()*{l_value})"
                    ws3.cell(row=r, column=6).value = f"=1/E{r}"
                ws3["H1"].value = "Average Conductivity"
                ws3["H2"].value = f"=AVERAGE(F2:F{last_row})"
            wb.save(path)
        except Exception as e:
            print(f"\n   ❌ Error in {os.path.basename(path)}: {e}")
        print_progress(i, total)
    print("\n   ✅ Step 1 Complete.")

def get_folder_data(excel, folder_path):
    f_name = os.path.basename(folder_path)
    m = re.search(r'(\d+)\s*%', f_name)
    rh_value = int(m.group(1)) if m else 0

    files_in_folder = [os.path.join(folder_path, f) for f in os.listdir(folder_path) 
                       if f.endswith(".xlsx") and not f.startswith("~$") and "Summary" not in f]

    avg_values = []
    for path in files_in_folder:
        try:
            wb = excel.Workbooks.Open(os.path.abspath(path))
            val = wb.Sheets(3).Range("H2").Value
            if val is not None:
                avg_values.append(val)
            wb.Close(False)
        except: pass
    
    if avg_values:
        folder_avg = sum(avg_values) / len(avg_values)
        return (rh_value, folder_avg)
    return None

# --- [3. 메인 실행 루프] ---
def main():
    while True:
        root_input = input("\n📂 Enter root folder: ").strip().strip('"')
        if not os.path.isdir(root_input):
            print("❌ Invalid path."); continue
        try:
            l_val = float(input("✍️ Enter l (cm): "))
        except:
            print("❌ Enter a number."); continue
        
        start = time.time()
        apply_formulas_to_all_files(root_input, l_val)
        
        try:
            excel = win32.gencache.EnsureDispatch('Excel.Application')
        except AttributeError:
            excel = win32.Dispatch('Excel.Application')
            
        excel.Visible, excel.DisplayAlerts = False, False
        
        try:
            print(f"📊 Step 2: Collecting data and creating clean chart...")
            subs = [os.path.join(root_input, f) for f in os.listdir(root_input) 
                    if os.path.isdir(os.path.join(root_input, f))]
            
            summary_data = []
            for s in subs:
                res = get_folder_data(excel, s)
                if res: summary_data.append(res)
            
            summary_data.sort(key=lambda x: x[0])

            if summary_data:
                wb_out = excel.Workbooks.Add()
                ws_out = wb_out.ActiveSheet
                
                ws_out.Cells(1,1).Value = "Relative humidity (%)"
                ws_out.Cells(1,2).Value = "σ (S/cm)"
                ws_out.Rows(1).Font.Bold = True
                
                for row_idx, (rh, sigma) in enumerate(summary_data, 2):
                    ws_out.Cells(row_idx, 1).Value = rh
                    ws_out.Cells(row_idx, 2).Value = sigma
                
                last_row = len(summary_data) + 1
                chart_range = ws_out.Range(ws_out.Cells(1,1), ws_out.Cells(last_row, 2))
                chart_obj = ws_out.ChartObjects().Add(150, 20, 450, 300)
                chart = chart_obj.Chart
                chart.ChartType = 74 
                chart.SetSourceData(chart_range)
                chart.HasLegend = False

                # 서식 설정
                if chart.SeriesCollection().Count > 0:
                    series = chart.SeriesCollection(1)
                    accent_color = 8544277 # RGB(21, 96, 130)
                    series.MarkerStyle = 8 
                    series.MarkerSize = 5
                    series.MarkerBackgroundColor = accent_color
                    series.MarkerForegroundColor = accent_color
                    series.Format.Line.Weight = 1.5
                    series.Format.Line.ForeColor.RGB = accent_color
                
                # --- [눈금선 제거 루프 강화] ---
                for ax_type in [1, 2]: # 1: X축(Category), 2: Y축(Value)
                    ax = chart.Axes(ax_type)
                    ax.HasMajorGridlines = False # 주 눈금선 끄기
                    ax.HasMinorGridlines = False # 보조 눈금선 끄기
                    ax.HasTitle = True
                    if ax_type == 1:
                        ax.AxisTitle.Text = "Relative humidity (%)"
                    else:
                        ax.AxisTitle.Text = "σ (S/cm)"
                    ax.AxisTitle.Font.Size = 15
                    ax.AxisTitle.Font.Bold = True

                ws_out.Columns("A:B").AutoFit()
                ws_out.Columns("A:B").HorizontalAlignment = -4108

                save_name = f"Total_Summary_{int(time.time())}.xlsx"
                save_path = os.path.join(root_input, save_name)
                wb_out.SaveAs(os.path.abspath(save_path))
                wb_out.Close()
                print(f"   ✅ Integrated Summary created: {save_name}")

        finally:
            excel.Quit()
        
        print(f"\n✨ Done! ({time.time()-start:.1f}s)")
        print("Press [Ctrl+F] to continue, any other key to exit.")
        if msvcrt.getch() != b'\x06': break

if __name__ == "__main__":
    main()